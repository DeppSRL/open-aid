# coding=utf-8
__author__ = 'stefano'
import logging
from pprint import pprint
from datetime import datetime
from django.core.exceptions import ObjectDoesNotExist
from openpyxl import load_workbook, Workbook
from django.core.management.base import BaseCommand
from openaid.projects.models import Project, Activity, Initiative, NewProject


class Command(BaseCommand):

    help = 'export file with initiatives for check'
    logger = logging.getLogger('openaid')



    def handle(self, *args, **options):
        verbosity = options['verbosity']
        if verbosity == '0':
            self.logger.setLevel(logging.ERROR)
        elif verbosity == '1':
            self.logger.setLevel(logging.WARNING)
        elif verbosity == '2':
            self.logger.setLevel(logging.INFO)
        elif verbosity == '3':
            self.logger.setLevel(logging.DEBUG)

        output_filename= 'initiatives.xlsx'
        initiative_fields = ['code', 'pk', 'title_en', 'recipient_temp','total_project_costs','last_update_temp','updated_at']
        workbook = Workbook()
        ws_output = workbook.create_sheet(index=0, title='sheet')
        self.logger.info(u"start")

        # append headers to output file
        ws_output.append(initiative_fields)

        for init in Initiative.objects.all().order_by('code'):

            row = []

            # gets data from initiative
            for f in initiative_fields:
                value = getattr(init, f)
                if f == 'recipient_temp' and value is not None:
                    row.append(value.name)
                else:
                    row.append(value)

            ws_output.append(row)

        # save output file
        workbook.save(output_filename)
        self.logger.info(u"finish")